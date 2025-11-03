import streamlit as st
import pandas as pd
import numpy as np
import re
import io
from openpyxl.styles import Font
from datetime import datetime, date

# --- APP CONFIGURATION ----------------------------------------------------

st.set_page_config(
    page_title="Dentist Scheduling Tool",
    page_icon="evenflow_ai_logo.svg",
    layout="wide"
)

# custom CSS
st.markdown("""
<style>
    /* Main container styling (existing) */
    .main .block-container {
        padding-top: 2rem;
        padding-bottom: 2rem;
    }
    /* Expander styling (existing) */
    .st-expander {
        border-radius: 10px;
        border: 1px solid #e6e6e6;
    }
    .st-expander header {
        font-size: 1.25rem;
        font-weight: bold;
    }
    /* Footer styling (existing) */
    .footer {
        position: fixed;
        left: 0;
        bottom: 0;
        width: 100%;
        background-color: #f1f1f1;
        color: #555;
        text-align: center;
        padding: 10px;
        font-size: 0.9rem;
        z-index: 99;
    }
    /* --- NEW: BLUE THEME FOR INTERACTIVE WIDGETS --- */
    /* Primary Buttons ('Load Files', 'Generate Report') & Download Button */
    .stButton > button, .stDownloadButton > button {
        background-color: #3366cc;
        color: white;
        border: 1px solid #3366cc;
    }
    /* Hover and active states for buttons */
    .stButton > button:hover, .stDownloadButton > button:hover {
        background-color: #2659b3; /* Slightly darker blue */
        border-color: #2659b3;
        color: white;
    }
    .stButton > button:active, .stDownloadButton > button:active {
        background-color: #1a4c99; /* Even darker blue for click */
        border-color: #1a4c99;
        color: white;
    }
    /* Slider ('Months Since Last RO') */
    div[data-testid="stThumbValue"] {
        color: #3366cc;
    }
    div[data-testid="stFilledTrack"] {
        background: #3366cc;
    }
    /* Multiselect - Tags for selected items */
    div[data-baseweb="tag"] {
         background-color: #e0e9ff; /* A lighter blue for the tags */
         color: #1a3366; /* Darker blue text for readability */
     }
    /* Dropdown menu - hover and selected item highlight */
    li[data-baseweb="menu"] > div[aria-selected="true"],
    li[data-baseweb="menu"] > div:hover {
        background-color: #e0e9ff; /* Light blue for selection/hover in dropdown */
        color: #1c1c1c; /* Black text for readability */
    }
</style>
""", unsafe_allow_html=True)


# Function to intelligently load data and find the header
@st.cache_data
def load_data(uploaded_file, required_columns):
    """
    Intelligently loads a CSV or Excel file by searching for the header row
    within the first 20 rows of the file.
    """
    if uploaded_file is None:
        return None
    try:
        file_content = io.BytesIO(uploaded_file.getvalue())

        def check_columns(df, req_cols):
            df.columns = [str(col).strip() for col in df.columns]
            return all(col in df.columns for col in req_cols)

        if uploaded_file.name.endswith('.csv'):
            for i in range(20):
                try:
                    file_content.seek(0)
                    df = pd.read_csv(file_content, header=i, low_memory=False, on_bad_lines='skip')
                    if check_columns(df, required_columns):
                        st.info(f"Header found on row {i+1} in '{uploaded_file.name}'.")
                        return df
                except Exception:
                    continue
        elif uploaded_file.name.endswith(('.xls', '.xlsx')):
            for i in range(20):
                try:
                    df = pd.read_excel(file_content, header=i)
                    if check_columns(df, required_columns):
                        st.info(f"Header found on row {i+1} in '{uploaded_file.name}'.")
                        return df
                except Exception:
                    continue

        st.error(f"Could not find a valid header with all required columns {required_columns} in the first 20 rows of '{uploaded_file.name}'. Please check the file format.")
        return None
    except Exception as e:
        st.error(f"An error occurred while loading or parsing {uploaded_file.name}: {e}")
        return None


# The cleaning functions are adapted to work with the dataframe in memory
def clean_appointments(df):
    df = df.rename(columns={'sc_name': 'service_center', 'Op Code': 'op_code'})
    df = df.rename(columns=lambda x: x.strip())
    
    email_col_found = any('email' in str(col).lower() for col in df.columns)
    
    if not email_col_found:
        for col in df.columns:
            if str(col).lower().startswith('unnamed'):
                df.rename(columns={col: 'Email'}, inplace=True)
                st.info(f"An unnamed column was identified as the email column and renamed to 'Email'.")
                break
    else:
        for col in df.columns:
            if 'email' in str(col).lower():
                df.rename(columns={col: 'Email'}, inplace=True)
                break
                
    if 'VIN' not in df.columns: raise ValueError("VIN column missing from Appointments file.")
    df['VIN'] = df['VIN'].astype(str).str.strip().str.upper()
    for col in ['service_center', 'reporting_status', 'op_code']:
        if col in df.columns: df[col] = df[col].astype(str).str.strip()
    df['Created Date'] = pd.to_datetime(df['Created Date'], errors='coerce')
    df['Planned Date'] = pd.to_datetime(df['Planned Date'], errors='coerce')
    df['customer_id'] = pd.to_numeric(df.get('customer_id', '0').astype(str).str.replace(',', ''), errors='coerce')
    if 'Vehicle' in df.columns:
        df['Vehicle'] = df['Vehicle'].astype(str)
        df['Vehicle_Year'] = df['Vehicle'].str.extract(r'(\d{4})', expand=False)
        df['Vehicle_Year'] = pd.to_numeric(df['Vehicle_Year'], errors='coerce')
        df['Vehicle_Brand'] = df['Vehicle'].str.split().str[0].str.upper().str.strip()
    df.dropna(subset=['Planned Date', 'customer_id', 'VIN', 'Vehicle_Year', 'service_center', 'op_code', 'Vehicle'], inplace=True)
    return df

def clean_ro(df):
    df = df.rename(columns={'vehicle_vin': 'VIN'})
    if 'VIN' not in df.columns: raise ValueError("vehicle_vin column missing from Repair Orders file.")
    df['VIN'] = df['VIN'].astype(str).str.strip().str.upper()
    df['open_date'] = pd.to_datetime(df.get('open_date'), errors='coerce')
    df['closed_date'] = pd.to_datetime(df.get('closed_date'), errors='coerce')
    df['customer_id'] = pd.to_numeric(df.get('customer_id', '0').astype(str).str.replace(',', ''), errors='coerce')
    df.dropna(subset=['open_date', 'closed_date', 'customer_id', 'VIN'], inplace=True)
    return df

# Core filtering logic
def apply_base_filters(appts_df, config):
    df = appts_df.copy()
    if config.get("PLANNED_DATE_START"): df = df[df['Planned Date'] >= config['PLANNED_DATE_START']]
    if config.get('EXCLUDE_VEHICLE_KEYWORDS'):
        regex_pattern = '|'.join(config['EXCLUDE_VEHICLE_KEYWORDS'])
        df = df[~df['Vehicle'].str.contains(regex_pattern, case=False, na=False)]
    mask = (df['reporting_status'].isin(config['REPORTING_STATUS'])) & (df['Vehicle_Brand'].isin(config['BRANDS']))
    df = df[mask]
    return df

def apply_lapsed_vehicle_filters(appts_df, ro_df, config, progress_callback):
    progress_callback(0.1, "Identifying lapsed vehicles...")
    def get_target_vins(all_appts_df, ro_df, months_ago, report_date):
        cutoff_date = report_date - pd.DateOffset(months=months_ago)
        latest_ro_dates = ro_df.groupby('VIN')['closed_date'].max()
        vins_with_old_ros = latest_ro_dates[latest_ro_dates < cutoff_date].index
        vins_with_recent_appts = all_appts_df[all_appts_df['Planned Date'] >= cutoff_date]['VIN'].unique()
        target_vins = np.setdiff1d(vins_with_old_ros, vins_with_recent_appts)
        return target_vins, cutoff_date

    report_date = pd.to_datetime(config['REPORTING_DATE']).normalize()
    target_vins, cutoff_date = get_target_vins(appts_df, ro_df, config['MONTHS_SINCE_LAST_RO'], report_date)
    progress_callback(0.3, "Applying base filters..."); filtered_df = apply_base_filters(appts_df, config)
    progress_callback(0.5, "Applying lapsed vehicle logic..."); filtered_df = filtered_df[filtered_df['VIN'].isin(target_vins)]; filtered_df = filtered_df[filtered_df['Planned Date'] < cutoff_date]
    progress_callback(0.7, "Applying specific filters...");

    # MODIFIED: Filter by selected Service Centers first (mandatory). Then, optionally filter by specified Op Codes.
    if config.get("SELECTED_SERVICE_CENTERS"):
        filtered_df = filtered_df[filtered_df['service_center'].isin(config['SELECTED_SERVICE_CENTERS'])]

    if config.get("SERVICE_CENTER_OP_CODES"): # This dict is only populated if user selects op codes
        op_code_masks = [((filtered_df['service_center'] == center) & (filtered_df['op_code'].isin(op_codes))) for center, op_codes in config["SERVICE_CENTER_OP_CODES"].items()]
        if op_code_masks: filtered_df = filtered_df[np.logical_or.reduce(op_code_masks)]

    if config.get('MODEL_YEARS'): filtered_df = filtered_df[filtered_df['Vehicle_Year'].isin(config['MODEL_YEARS'])]
    return filtered_df

def apply_general_summary_filters(appts_df, config):
    summary_df = apply_base_filters(appts_df, config)
    # MODIFIED: Use the master list of selected service centers for summary filtering as well.
    if config.get("SELECTED_SERVICE_CENTERS"):
        summary_df = summary_df[summary_df['service_center'].isin(config['SELECTED_SERVICE_CENTERS'])]
    return summary_df

# Analysis and report generation logic
def calculate_stats(df, column_name):
    if df.empty or column_name not in df.columns or df[column_name].isnull().all(): return pd.DataFrame()
    counts = df[column_name].value_counts(); percentages = df[column_name].value_counts(normalize=True) * 100
    stats_df = pd.DataFrame({'Count': counts, 'Percentage': percentages.round(2)})
    return stats_df.reindex(counts.index).reset_index().rename(columns={'index': 'Category'})
def analyze_booking_days(df):
    df_copy = df.dropna(subset=['Created Date']).copy();
    if df_copy.empty: return pd.DataFrame()
    df_copy['Advanced Booking Days Raw'] = (df_copy['Planned Date'] - df_copy['Created Date']).dt.days; bins_days = [-float('inf'), 3, 5, 7, float('inf')]; labels_days = ['0-3 Days', '4-5 Days', '6-7 Days', '8+ Days']; df_copy['Booking_Day_Category'] = pd.cut(df_copy['Advanced Booking Days Raw'], bins=bins_days, labels=labels_days); return calculate_stats(df_copy, 'Booking_Day_Category')
def analyze_time_slots(df):
    df_copy = df.copy(); bins_time = [0, 7, 9, 11, 13, 15, 17, 24]; labels_time = ['Before 7am', '7am-9am', '9am-11am', '11am-1pm', '1pm-3pm', '3pm-5pm', 'After 5pm']; df_copy['Time_Slot'] = pd.cut(df_copy['Planned Date'].dt.hour, bins=bins_time, labels=labels_time, right=False, ordered=True); df_copy['Time_Slot'] = pd.Categorical(df_copy['Time_Slot'], categories=labels_time, ordered=True); return calculate_stats(df_copy, 'Time_Slot')
def analyze_day_of_week(df):
    df_copy = df.copy(); day_order = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]; df_copy['Day_of_Week'] = pd.Categorical(df_copy['Planned Date'].dt.day_name(), categories=day_order, ordered=True); weekday_df = df_copy[df_copy['Day_of_Week'].isin(day_order[:5])]; return calculate_stats(weekday_df, 'Day_of_Week')
def add_detailed_analysis_columns(df):
    df_out = df.copy(); df_out['Advanced Booking Days Raw'] = (df_out['Planned Date'] - df_out['Created Date']).dt.days; bins_days = [-float('inf'), 3, 5, 7, float('inf')]; labels_days = ['0-3 Days', '4-5 Days', '6-7 Days', '8+ Days']; df_out['Advanced Booking Days'] = pd.cut(df_out['Advanced Booking Days Raw'], bins=bins_days, labels=labels_days); bins_time = [0, 7, 9, 11, 13, 15, 17, 24]; labels_time = ['Before 7am', '7am-9am', '9am-11am', '11am-1pm', '1pm-3pm', '3pm-5pm', 'After 5pm']; df_out['Appointment Time Slot'] = pd.cut(df_out['Planned Date'].dt.hour, bins=bins_time, labels=labels_time, right=False, ordered=True); day_order = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]; df_out['Appointment Day of Week'] = pd.Categorical(df_out['Planned Date'].dt.day_name(), categories=day_order, ordered=True); return df_out

def write_summary_to_sheet(writer, sheet_name, summaries, total_records):
    worksheet = writer.sheets[sheet_name]; bold_font = Font(bold=True)
    worksheet.cell(row=1, column=1, value=f"Analysis for: {sheet_name.replace('_', ' ')}").font = bold_font
    worksheet.cell(row=2, column=1, value=f"Total Unique Vehicles Analyzed: {total_records}")
    if total_records == 0:
        worksheet.cell(row=4, column=1, value="No matching vehicle data found for this summary.")
        return
    start_row = 5
    for title, df in summaries.items():
        worksheet.cell(row=start_row, column=1, value=title).font = bold_font
        if not df.empty: df['Percentage'] = df['Percentage'].apply(lambda x: f"{x:.2f}%"); df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=start_row, header=True); start_row += len(df) + 3
        else: worksheet.cell(row=start_row + 1, column=1, value="No data available for this category."); start_row += 3

def sanitize_sheet_name(name):
    clean_name = name.replace("Walser ", "").replace(" of ", " "); clean_name = re.sub(r'[^a-zA-Z0-9_]', '', clean_name).strip('_'); return clean_name[:31]

def generate_excel_report(lapsed_df, summary_df, ro_df, config):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # MODIFIED: Aggregate lapsed_df to get unique VINs and a new "Used Op Codes" column
        if not lapsed_df.empty:
            # Group by VIN and customer_id to get all Op Codes for each unique vehicle appointment
            op_code_agg = lapsed_df.groupby(['VIN', 'customer_id'])['op_code'].apply(lambda x: ', '.join(sorted(x.astype(str).unique()))).reset_index(name='Used Op Codes')
            
            # Keep the rest of the data from the most recent appointment record for that vehicle
            main_data = lapsed_df.sort_values('Planned Date', ascending=False).drop_duplicates(subset=['VIN', 'customer_id'], keep='first')
            
            # Merge to create the final detailed dataframe, dropping the original 'op_code'
            lapsed_df_aggregated = pd.merge(main_data.drop('op_code', axis=1), op_code_agg, on=['VIN', 'customer_id'])
        else:
            lapsed_df_aggregated = pd.DataFrame()

        # 1. Detailed Sheet
        detailed_output_df = add_detailed_analysis_columns(lapsed_df_aggregated)
        if not ro_df.empty and 'VIN' in detailed_output_df.columns:
            last_ro_indices = ro_df.groupby('VIN')['closed_date'].idxmax()
            last_ro_info = ro_df.loc[last_ro_indices, ['VIN', 'open_date', 'closed_date']].rename(columns={'open_date': 'Last RO Open Date', 'closed_date': 'Last RO Closed Date'})
            detailed_output_df = pd.merge(detailed_output_df, last_ro_info, on='VIN', how='left')

        detailed_output_df.to_excel(writer, sheet_name='Detailed_Data', index=False)
        
        # 2. Summary Sheets
        summary_df_unique_vin = summary_df.sort_values('Planned Date', ascending=False).drop_duplicates(subset=['VIN'], keep='first')
        
        summaries_overall = {
            "Advanced Booking Days": analyze_booking_days(summary_df_unique_vin), 
            "Appointment Time Slots": analyze_time_slots(summary_df_unique_vin), 
            "Appointment Day of Week (Mon-Fri)": analyze_day_of_week(summary_df_unique_vin)
        }
        pd.DataFrame().to_excel(writer, sheet_name='Overall_Summary')
        write_summary_to_sheet(writer, 'Overall_Summary', summaries_overall, len(summary_df_unique_vin))
        
        if config.get("SELECTED_SERVICE_CENTERS"):
            for center in config["SELECTED_SERVICE_CENTERS"]:
                df_sc_summary = summary_df[summary_df['service_center'] == center]
                
                df_sc_summary_unique_vin = df_sc_summary.sort_values('Planned Date', ascending=False).drop_duplicates(subset=['VIN'], keep='first')
                
                summaries_sc = {
                    "Advanced Booking Days": analyze_booking_days(df_sc_summary_unique_vin), 
                    "Appointment Time Slots": analyze_time_slots(df_sc_summary_unique_vin), 
                    "Appointment Day of Week (Mon-Fri)": analyze_day_of_week(df_sc_summary_unique_vin)
                }
                sheet_name = f"{sanitize_sheet_name(center)}_Summary"
                pd.DataFrame().to_excel(writer, sheet_name=sheet_name)
                write_summary_to_sheet(writer, sheet_name, summaries_sc, len(df_sc_summary_unique_vin))

    processed_data = output.getvalue()
    return processed_data

# Initialize session state for application flow
if 'files_ok' not in st.session_state:
    st.session_state.files_ok = False
    st.session_state.appts_df = None
    st.session_state.ro_df = None
    st.session_state.report_bytes = None

# Header
st.title("Dentist Scheduling Analysis Tool")
st.markdown("This tool analyses appointment and RO Lines data to generate insights and targeted customers list")
st.divider()

# --- 1. File Upload Section ---
st.header("1. Upload Data Files")
col1, col2 = st.columns(2)
with col1:
    appointments_file = st.file_uploader("Upload Appointments File", type=["csv", "xlsx"], help="Upload the file with appointment lines. The app automatically finds the header.")
with col2:
    repair_orders_file = st.file_uploader("Upload Repair Orders File", type=["csv", "xlsx"], help="Upload the file with repair order history. Required for lapsed vehicle analysis.")

if appointments_file and repair_orders_file:
    if st.button("Load and Validate Files", use_container_width=True):
        with st.spinner("Reading and validating files..."):
            required_appt_cols = ['VIN', 'Created Date', 'Planned Date', 'Vehicle', 'sc_name', 'Op Code', 'reporting_status']
            required_ro_cols = ['vehicle_vin', 'open_date', 'closed_date']
            
            raw_appts_df = load_data(appointments_file, required_appt_cols)
            raw_ro_df = load_data(repair_orders_file, required_ro_cols)

            if raw_appts_df is not None and raw_ro_df is not None:
                try:
                    st.session_state.appts_df = clean_appointments(raw_appts_df)
                    st.session_state.ro_df = clean_ro(raw_ro_df)
                    st.session_state.files_ok = True
                    st.success("Files successfully loaded and validated! You can now configure the analysis.")
                except Exception as e:
                    st.session_state.files_ok = False
                    st.error(f"Error during data cleaning: {e}")
            else:
                st.session_state.files_ok = False
                st.warning("Please check the error messages above and upload valid files.")

if st.session_state.files_ok:
    st.success("✅ Files Loaded Successfully. Configure your report below.")
    df_appts = st.session_state.appts_df

    # --- 2. Configuration Section ---
    with st.expander("2. Configure Analysis Filters", expanded=True):
        st.subheader("Base Filters (Applied to ALL Reports)")
        b_col1, b_col2 = st.columns(2)
        with b_col1:
            planned_date_start = st.date_input("Planned Date Start", value=date(2024, 1, 1), help="Required. Analysis will only include appointments on or after this date.")
        with b_col2:
            reporting_status = st.multiselect("Reporting Status", options=sorted(df_appts['reporting_status'].unique()), default=sorted(df_appts['reporting_status'].unique()), help="Select appointment statuses to include.")
        b_col3, b_col4 = st.columns(2)
        with b_col3:
            brands = st.multiselect("Vehicle Brands", options=sorted(df_appts['Vehicle_Brand'].unique()), default=sorted(df_appts['Vehicle_Brand'].unique()), help="Select vehicle brands to include.")
        with b_col4:
            exclude_keywords = st.multiselect("Exclude Vehicle Keywords", options=["EV", "ELECTRIC", "HYBRID", "PHEV"], help="Select keywords to exclude vehicles (e.g., EVs). Case-insensitive.")

        st.divider()
        st.subheader("Filters for DETAILED Lapsed Vehicle Report ONLY")
        d_col1, d_col2 = st.columns(2)
        with d_col1:
            reporting_date = st.date_input("Reporting Date", value=date.today(), help="Required. The reference 'as of' date for the lapsed calculation.")
        with d_col2:
            months_since_ro = st.slider("Months Since Last RO", min_value=1, max_value=24, value=7, help="How many months without service to be considered 'lapsed'.")

        model_years = st.multiselect("Model Years", options=sorted(df_appts['Vehicle_Year'].unique().astype(int), reverse=True), help="Select specific model years for the detailed report.")
        
        st.markdown("**Service Center & Op Code Mapping (Op Codes are optional)**")
        sc_options = sorted(df_appts['service_center'].unique())
        selected_sc = st.multiselect("Select Service Centers", options=sc_options, help="Choose the service centers for the detailed report and summaries.")
        
        service_center_op_codes = {}
        for center in selected_sc:
            c1, c2 = st.columns([2, 3])
            with c1: st.markdown(f"**{center}**")
            with c2:
                op_options_for_center = sorted(df_appts[df_appts['service_center'] == center]['op_code'].unique())
                selected_ops = st.multiselect(f"Select Op Codes for {center}", options=op_options_for_center, key=f"op_{center}", help=f"Optionally, choose op codes to filter for {center}. If none are selected, all op codes for this center will be considered.")
                if selected_ops: service_center_op_codes[center] = selected_ops

# --- 3. Run Analysis Section ---
if st.session_state.files_ok and 'df_appts' in locals():
    if st.button("Generate Report", type="primary", use_container_width=True):
        # MODIFIED: Op codes are now optional. The report only requires a service center to be selected.
        if not planned_date_start or not reporting_date or not reporting_status or not brands or not selected_sc:
            st.warning("Please ensure all required fields are filled: Dates, Status, Brands, and at least one Service Center.")
        else:
            config = {
                "PLANNED_DATE_START": pd.to_datetime(planned_date_start), "EXCLUDE_VEHICLE_KEYWORDS": exclude_keywords,
                "REPORTING_STATUS": reporting_status, "BRANDS": brands, "MONTHS_SINCE_LAST_RO": months_since_ro,
                "REPORTING_DATE": pd.to_datetime(reporting_date), "MODEL_YEARS": model_years, 
                "SELECTED_SERVICE_CENTERS": selected_sc,
                "SERVICE_CENTER_OP_CODES": service_center_op_codes,
            }

            progress_bar = st.progress(0, "Starting analysis...")
            try:
                def update_progress(value, text): progress_bar.progress(value, text=text)
                lapsed_df = apply_lapsed_vehicle_filters(st.session_state.appts_df, st.session_state.ro_df, config, update_progress)
                update_progress(0.8, "Filtering data for summaries...")
                summary_df = apply_general_summary_filters(st.session_state.appts_df, config)
                update_progress(0.9, "Generating Excel report...")
                st.session_state.report_bytes = generate_excel_report(lapsed_df, summary_df, st.session_state.ro_df, config)
                update_progress(1.0, "Analysis Complete!")
            except Exception as e:
                st.error(f"An error occurred during analysis: {e}")
                st.session_state.report_bytes = None

# --- 4. Download Section ---
if st.session_state.get('report_bytes'):
    st.divider()
    st.header("✅ Report Generated Successfully!")
    file_name = f"Dentist_Scheduling_Analysis_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    st.download_button(label="Download Excel Report", data=st.session_state.report_bytes, file_name=file_name,
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)

# --- 5. Footer ---
st.markdown("""
<div class="footer">
  <p>All Rights Reserved | An EvenFlow AI Tool  | V 2.0 | 2025</p>
</div>
""", unsafe_allow_html=True)